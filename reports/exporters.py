import csv
from io import StringIO, BytesIO
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from .widgets import WidgetFactory

class BaseExporter:
    """Abstract base class for all exporters"""
    def __init__(self, dashboard):
        self.dashboard = dashboard
    
    def export(self):
        """Must be implemented by subclasses"""
        raise NotImplementedError

class PDFExporter(BaseExporter):
    """Export dashboard to PDF"""
    def export(self):
        html = render_to_string('reports/dashboard_pdf.html', {
            'dashboard': self.dashboard,
            'widget_data': self.get_widget_data()
        })
        
        result = BytesIO()
        pdf = pisa.pisaDocument(StringIO(html), result)
        
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{self.dashboard.name}.pdf"'
            return response
        return HttpResponse('Error generating PDF', status=500)
    
    def get_widget_data(self):
        data = {}
        for widget in self.dashboard.widgets.all():
            widget_instance = WidgetFactory.create_widget(widget.data_source, widget.config, user=None)
            data[widget.id] = widget_instance.render()
        return data

class CSVExporter(BaseExporter):
    """Export dashboard data to CSV"""
    def export(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.dashboard.name}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Dashboard:', self.dashboard.name])
        writer.writerow([])
        
        for widget in self.dashboard.widgets.all():
            self.export_widget(writer, widget)
        
        return response
    
    def export_widget(self, writer, widget):
        widget_instance = WidgetFactory.create_widget(widget.data_source, widget.config, user=None)
        data = widget_instance.get_data()
        
        writer.writerow([f"Widget: {widget.title} ({widget.get_widget_type_display()})"])
        
        if not data:
            writer.writerow(["No data available"])
            writer.writerow([])
            return
        
        # Write headers
        headers = list(data[0].keys())
        writer.writerow(headers)
        
        # Write data rows
        for row in data:
            writer.writerow([str(row.get(key, '')) for key in headers])
        
        writer.writerow([])

class ExcelExporter(BaseExporter):
    """Export dashboard to Excel (using openpyxl)"""
    def export(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Report"
        
        # Add title
        ws['A1'] = self.dashboard.name
        ws['A1'].font = Font(bold=True, size=14)
        
        row_num = 3
        
        for widget in self.dashboard.widgets.all():
            widget_instance = WidgetFactory.create_widget(widget.data_source, widget.config, user=None)
            data = widget_instance.get_data()
            
            # Add widget title
            ws.cell(row=row_num, column=1, value=f"{widget.title} ({widget.get_widget_type_display()})")
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            row_num += 1
            
            if not data:
                ws.cell(row=row_num, column=1, value="No data available")
                row_num += 2
                continue
            
            # Add headers
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=header)
                ws.cell(row=row_num, column=col_num).font = Font(bold=True)
            
            row_num += 1
            
            # Add data
            for row in data:
                for col_num, key in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col_num, value=str(row.get(key, '')))
                row_num += 1
            
            row_num += 2  # Add spacing between widgets
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.dashboard.name}.xlsx"'
        wb.save(response)
        
        return response

class ExporterFactory:
    """Factory class to create appropriate exporter"""
    EXPORTER_MAP = {
        'pdf': PDFExporter,
        'csv': CSVExporter,
        'excel': ExcelExporter,
    }
    
    @classmethod
    def create_exporter(cls, format, dashboard):
        exporter_class = cls.EXPORTER_MAP.get(format.lower())
        if not exporter_class:
            raise ValueError(f"Unsupported export format: {format}")
        return exporter_class(dashboard)
    
    @classmethod
    def get_available_formats(cls):
        return list(cls.EXPORTER_MAP.keys())